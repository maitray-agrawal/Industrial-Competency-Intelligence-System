import os
import re
import atexit
import hashlib
import uuid as _uuid
from datetime import datetime
from flask import Flask, render_template, request, Response, redirect, url_for, flash, jsonify, session as flask_session, send_file, abort
from werkzeug.utils import secure_filename
from openpyxl import load_workbook

from database import SessionLocal, engine, init_db
from models import (
    Shop, Station, Process, Operation, Skill, Tool,
    Topic, Subtopic, StagingData, CompetencyMap,
    SkillOperationMap, TopicSkillMap, ToolStationMap, UploadedFile,
    SkillStationMap, StationOperationMap,
    GraphEntity, GraphRelationship,
    ShopWISDocument, StationWISDocument,
    ShopWISWorkbook, StationWISSheet,
    ShopPPEWorkbook, StationPPESheet,
)
from sqlalchemy import Integer as _SAInteger
from search_engine import SearchAPI, SearchIndexer
from data_engine import IngestionPipeline, SHOP_ALIASES
from heuristic_engine import KnowledgeMapper
from competency_engine import CompetencyEngine
from logger import get_logger
from api_contracts import build_unified_search_entity
from sqlalchemy import text, func
from sqlalchemy.orm import joinedload

logger = get_logger("FlaskApp")

# ---------------------------------------------------------------------------
# App Configuration
# ---------------------------------------------------------------------------
app = Flask(__name__)
app.secret_key = os.environ.get("IIK_SECRET_KEY", "industrial-offline-secret-2025")
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB upload limit

UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
ALLOWED_EXTENSIONS = {".xlsx", ".xls"}
ALLOWED_WIS_EXTENSIONS = {".ppt", ".pptx"}
ALLOWED_WORKBOOK_EXTENSIONS = {".xlsx", ".xlsm"}

DEBUG_MODE = os.environ.get("FLASK_ENV", "production").lower() == "development"

def allowed_file(filename: str) -> bool:
    """Return True only if the file has an allowed Excel extension."""
    ext = os.path.splitext(filename.lower())[1]
    return ext in ALLOWED_EXTENSIONS


def allowed_wis_file(filename: str) -> bool:
    """Return True only if the file has an allowed PowerPoint extension."""
    ext = os.path.splitext(filename.lower())[1]
    return ext in ALLOWED_WIS_EXTENSIONS


def allowed_workbook_file(filename: str) -> bool:
    """Return True only if the file has an allowed spreadsheet workbook extension."""
    ext = os.path.splitext(filename.lower())[1]
    return ext in ALLOWED_WORKBOOK_EXTENSIONS


def _normalize_sheet_token(value: str | None) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def _match_station_for_sheet(sheet_name: str, stations: list[Station]) -> Station | None:
    sheet_key = _normalize_sheet_token(sheet_name)
    if not sheet_key:
        return None

    for station in stations:
        candidates = [station.station_code, station.raw_station_id, station.name]
        for candidate in candidates:
            if _normalize_sheet_token(candidate) == sheet_key:
                return station
            if sheet_key and _normalize_sheet_token(candidate) and sheet_key in _normalize_sheet_token(candidate):
                return station
    return None


def _normalize_shop_lookup_value(value: str | None) -> str:
    if not value:
        return ""
    return re.sub(r"[^A-Za-z0-9]+", "_", str(value).strip()).strip("_").upper()


def _resolve_shop_record(db, shop_code: str | None) -> Shop | None:
    raw = (shop_code or "").strip()
    candidates: set[str] = set()
    if raw:
        candidates.update({raw, raw.upper(), _normalize_shop_lookup_value(raw)})
        candidates.add(raw.replace(" ", "_"))
        candidates.add(raw.replace("_", " "))

    for candidate in sorted(candidates):
        if not candidate:
            continue
        shop = db.query(Shop).filter_by(shop_code=candidate).first()
        if shop:
            return shop

    for candidate in sorted(candidates):
        if not candidate:
            continue
        shop = db.query(Shop).filter(Shop.name.ilike(f"%{candidate}%")) .first()
        if shop:
            return shop
    return None


def _read_sheet_preview_rows(file_path: str, sheet_name: str, limit: int = 10) -> list[list[str]]:
    if not file_path or not os.path.exists(file_path):
        return []
    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            return []
        worksheet = workbook[sheet_name]
        rows: list[list[str]] = []
        for row in worksheet.iter_rows(min_row=1, max_row=min(limit, worksheet.max_row), values_only=True):
            rows.append(["" if cell is None else str(cell) for cell in row])
        workbook.close()
        return rows
    except Exception as exc:
        logger.warning(f"Unable to read sheet preview for {sheet_name}: {exc}")
        return []


def _get_workbook_mimetype(file_path: str) -> str:
    ext = os.path.splitext(file_path.lower())[1]
    if ext == ".xlsx":
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if ext == ".xlsm":
        return "application/vnd.ms-excel.sheet.macroEnabled.12"
    if ext == ".xls":
        return "application/vnd.ms-excel"
    return "application/octet-stream"


def _ensure_workbook_versioning_schema() -> None:
    try:
        with engine.begin() as conn:
            for table_name in ("shop_wis_workbooks", "shop_ppe_workbooks"):
                columns = {row[1] for row in conn.execute(text(f"PRAGMA table_info('{table_name}')")).fetchall()}
                if "version_number" not in columns:
                    conn.execute(text(f"ALTER TABLE {table_name} ADD COLUMN version_number INTEGER NOT NULL DEFAULT 1"))
                if "active" not in columns:
                    conn.execute(text(f"ALTER TABLE {table_name} ADD COLUMN active BOOLEAN NOT NULL DEFAULT 1"))
                if "archived_at" not in columns:
                    conn.execute(text(f"ALTER TABLE {table_name} ADD COLUMN archived_at DATETIME"))
                if "change_summary" not in columns:
                    conn.execute(text(f"ALTER TABLE {table_name} ADD COLUMN change_summary TEXT"))
                conn.execute(text(f"CREATE INDEX IF NOT EXISTS ix_{table_name}_active ON {table_name} (shop_id, active)"))
            conn.commit()
    except Exception as exc:
        logger.warning(f"Workbook versioning schema check skipped: {exc}")


def _get_next_workbook_version(db, workbook_cls, shop_id: int) -> int:
    latest = db.query(func.max(workbook_cls.version_number)).filter(workbook_cls.shop_id == shop_id).scalar() or 0
    return int(latest) + 1


def _build_workbook_change_summary(explicit_summary: str | None, workbook_kind: str, previous_active_workbook) -> str:
    if explicit_summary and explicit_summary.strip():
        return explicit_summary.strip()
    if previous_active_workbook:
        return f"Revised {workbook_kind.upper()} workbook upload; previous active workbook archived."
    return f"Initial {workbook_kind.upper()} workbook upload."


def _archive_previous_active_workbook(db, workbook_cls, shop_id: int, new_version_number: int) -> None:
    previous_active_workbooks = (
        db.query(workbook_cls)
        .filter(workbook_cls.shop_id == shop_id, workbook_cls.active.is_(True))
        .order_by(workbook_cls.uploaded_at.desc(), workbook_cls.id.desc())
        .all()
    )
    for previous_active in previous_active_workbooks:
        previous_active.active = False
        previous_active.archived_at = datetime.utcnow()
        if not previous_active.change_summary:
            previous_active.change_summary = f"Archived when version {new_version_number} was uploaded."
        else:
            previous_active.change_summary = f"{previous_active.change_summary} | Archived when version {new_version_number} was uploaded."


def _merge_sheet_mapping(mapping_row, matched_station, sheet_index: int | None, previous_mapping_row=None) -> None:
    if matched_station is not None:
        mapping_row.station_id = matched_station.id
        mapping_row.match_status = "matched"
    elif getattr(mapping_row, "station_id", None) is None and previous_mapping_row is not None and getattr(previous_mapping_row, "station_id", None) is not None:
        mapping_row.station_id = previous_mapping_row.station_id
        mapping_row.match_status = previous_mapping_row.match_status or "unmatched"
    elif getattr(mapping_row, "match_status", None) in (None, ""):
        mapping_row.match_status = "unmatched"

    if sheet_index is not None:
        mapping_row.sheet_index = sheet_index
    elif getattr(mapping_row, "sheet_index", None) is None and previous_mapping_row is not None:
        mapping_row.sheet_index = previous_mapping_row.sheet_index

    if getattr(mapping_row, "match_status", None) in (None, ""):
        mapping_row.match_status = previous_mapping_row.match_status or "unmatched" if previous_mapping_row is not None else "unmatched"


# ---------------------------------------------------------------------------
# Auth — HTTP Basic Auth for admin panel
# ---------------------------------------------------------------------------
ADMIN_USER = os.environ.get("IIK_ADMIN_USER", "admin")
ADMIN_PASS = os.environ.get("IIK_ADMIN_PASS", "tata@1945")


def check_auth(username, password):
    return username == ADMIN_USER and password == ADMIN_PASS


def authenticate():
    return Response(
        "Authentication required.\n",
        401,
        {"WWW-Authenticate": 'Basic realm="IIK Admin"'},
    )


def requires_auth(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.authorization
        if not auth or not check_auth(auth.username, auth.password):
            return authenticate()
        return f(*args, **kwargs)
    return decorated


# ---------------------------------------------------------------------------
# Error Handlers
# ---------------------------------------------------------------------------
@app.errorhandler(404)
def not_found(e):
    return render_template("error.html", code=404, message="Page not found."), 404


@app.errorhandler(500)
def server_error(e):
    logger.error(f"500 error: {e}")
    return render_template("error.html", code=500, message="Internal server error."), 500


@app.errorhandler(413)
def too_large(e):
    flash("File too large. Maximum upload size is 16 MB.", "danger")
    return redirect(url_for("admin"))


# ---------------------------------------------------------------------------
# Helper — system stats for dashboard
# ---------------------------------------------------------------------------
def _get_system_stats() -> dict:
    with SessionLocal() as db:
        return {
            # Exclude the virtual/wrapper WELD_SHOP from visible shop counts
            "shops":      db.query(Shop).filter(Shop.shop_code != "WELD_SHOP").count(),
            "stations":   db.query(Station).count(),
            "processes":  db.query(Process).count(),
            "operations": db.query(Operation).count(),
            "skills":     db.query(Skill).count(),
            "tools":      db.query(Tool).count(),
            "topics":     db.query(Topic).count(),
            "subtopics":  db.query(Subtopic).count(),
            "mappings":   db.query(CompetencyMap).count(),
        }


# ===========================================================================
# ROUTES — Public
# ===========================================================================

@app.route("/")
def index():
    stats = _get_system_stats()
    competency_summary = CompetencyEngine.get_all_stations_summary()
    return render_template("index.html", stats=stats, competency_summary=competency_summary)


@app.route("/search")
def search():
    q = request.args.get("q", "").strip()
    entity_filter = request.args.get("type", None)
    results = []
    if q:
        results = SearchAPI.search(q, limit=20, entity_filter=entity_filter or None)
    return render_template(
        "search.html", query=q, results=results, entity_filter=entity_filter
    )


@app.route("/station/<int:station_id>")
def station_detail(station_id):
    profile = CompetencyEngine.get_station_knowledge_profile(station_id)
    if "error" in profile:
        flash(profile["error"], "danger")
        return redirect(url_for("index"))

    # --- shop_code for breadcrumb ---
    shop_code = None
    with SessionLocal() as _s:
        _stn = _s.get(Station, station_id)
        if _stn and _stn.shop:
            shop_code = _stn.shop.shop_code

    # --- skills_summary: unique skills with coverage status ---
    readiness      = profile.get("readiness", {})
    missing_names  = {s["skill_name"] for s in readiness.get("missing_skills", [])}
    seen_sk        = set()
    skills_summary = []
    for proc in profile.get("processes", []):
        for op in proc.get("operations", []):
            for sk in op.get("skills", []):
                if sk not in seen_sk:
                    seen_sk.add(sk)
                    skills_summary.append({"name": sk, "covered": sk not in missing_names})
    skills_summary.sort(key=lambda x: (not x["covered"], x["name"]))

    # --- graph nodes + edges for vis-network ---
    stn_nid    = f"stn_{station_id}"
    graph_nodes = [{"id": stn_nid, "label": profile.get("station_code", "STN"),
                    "group": "station", "size": 32}]
    graph_edges = []

    for i, proc in enumerate(profile.get("processes", [])):
        pid   = f"proc_{i}"
        label = proc["name"][:24] + ("\u2026" if len(proc["name"]) > 24 else "")
        graph_nodes.append({"id": pid, "label": label, "group": "process"})
        graph_edges.append({"from": stn_nid, "to": pid})

    for sk in skills_summary[:14]:
        sid   = f"sk_{hash(sk['name']) & 0xFFFFFF}"
        label = sk["name"][:22] + ("\u2026" if len(sk["name"]) > 22 else "")
        graph_nodes.append({"id": sid, "label": label,
                             "group": "skill_ok" if sk["covered"] else "skill_gap"})
        graph_edges.append({"from": stn_nid, "to": sid})

    for topic in readiness.get("covered_topics", [])[:7]:
        tid   = f"topic_{topic['topic_id']}"
        label = topic["topic_title"][:24] + ("\u2026" if len(topic["topic_title"]) > 24 else "")
        graph_nodes.append({"id": tid, "label": label, "group": "topic"})
        graph_edges.append({"from": tid, "to": stn_nid})

    # --- WIS/PPE documents for this station ---
    wis_docs = None
    wis_sheets = []
    ppe_sheets = []
    is_admin = False
    with SessionLocal() as _s:
        wis_docs = _s.query(StationWISDocument).filter_by(station_id=station_id).order_by(StationWISDocument.uploaded_at.desc()).all()
        wis_sheets = (
            _s.query(StationWISSheet)
            .join(ShopWISWorkbook)
            .options(joinedload(StationWISSheet.workbook), joinedload(StationWISSheet.station))
            .filter(StationWISSheet.station_id == station_id, ShopWISWorkbook.active.is_(True))
            .order_by(StationWISSheet.uploaded_at.desc(), StationWISSheet.sheet_index)
            .all()
        )
        ppe_sheets = (
            _s.query(StationPPESheet)
            .join(ShopPPEWorkbook)
            .options(joinedload(StationPPESheet.workbook), joinedload(StationPPESheet.station))
            .filter(StationPPESheet.station_id == station_id, ShopPPEWorkbook.active.is_(True))
            .order_by(StationPPESheet.uploaded_at.desc(), StationPPESheet.sheet_index)
            .all()
        )
        is_admin = request.authorization and check_auth(request.authorization.username, request.authorization.password)

    wis_sheet_views = []
    for sheet in wis_sheets:
        wis_sheet_views.append({"sheet": sheet})

    ppe_sheet_views = []
    for sheet in ppe_sheets:
        preview_rows = []
        if sheet.workbook and sheet.workbook.file_path and os.path.exists(sheet.workbook.file_path):
            preview_rows = _read_sheet_preview_rows(sheet.workbook.file_path, sheet.sheet_name)
        ppe_sheet_views.append({"sheet": sheet, "preview_rows": preview_rows})

    return render_template(
        "station.html",
        profile        = profile,
        shop_code      = shop_code,
        skills_summary = skills_summary,
        graph_nodes    = graph_nodes,
        graph_edges    = graph_edges,
        wis_docs       = wis_docs,
        wis_sheet_views = wis_sheet_views,
        ppe_sheet_views = ppe_sheet_views,
        is_admin       = is_admin,
    )


@app.route("/wis-sheet/<int:sheet_id>")
def open_wis_sheet(sheet_id):
    with SessionLocal() as _s:
        sheet = _s.get(StationWISSheet, sheet_id)
        if not sheet or not sheet.workbook:
            abort(404)
        workbook = sheet.workbook
        if not workbook.file_path or not os.path.exists(workbook.file_path):
            abort(404)
        return send_file(
            workbook.file_path,
            mimetype=_get_workbook_mimetype(workbook.file_path),
            as_attachment=False,
            download_name=workbook.file_name or os.path.basename(workbook.file_path),
        )


@app.route("/graph")
def graph():
    stats = _get_system_stats()
    return render_template("graph.html", stats=stats)

# ---------------------------------------------------------------------------
# Helper — parse raw_station_id into Line / Zone / Station segments
# ---------------------------------------------------------------------------
def _fmt_zone(zone_raw: str) -> str:
    """Format a raw zone value (e.g. '1.0', '2', 'Zone 3') into a clean integer string."""
    if not zone_raw:
        return ""
    # Strip trailing .0 from float-like strings (e.g. '1.0' → '1')
    try:
        return str(int(float(zone_raw)))
    except (ValueError, TypeError):
        return zone_raw.strip()


def _parse_station_hierarchy(stations, shop_code: str | None = None):
    """
    Accepts a list of Station ORM objects (already ordered by row_order from the
    query) and returns a structured hierarchy that PRESERVES Excel insertion order.
    Grouping adapts to `shop_code`. Top-level entries returned have the shape
    {label, zones: [{label, zone_badge, stations: [...]}, ...]}. The parser
    selects the most-relevant non-empty fields per shop type and never
    creates empty hierarchy levels.
    """
    shop_map = {
        "TCF_1":       ["cell", "line", "zone_no"],
        "ENGINE_SHOP": ["cell", "line", "zone_no"],
        "TRANSAXLE_SHOP": ["cell", "line", "zone_no"],
        "PRESS_SHOP": ["cell", "line", "zone_no"],
        "JLR":        ["cell", "line", "zone_no"],
        "EV_SHOP":     ["cell", "line", "zone_no"],
        "X4_BIW":      ["cell", "line", "zone_no"],
        "X1_BIW":      ["line", "zone_no"],
        "PAINT_SHOP":  ["line", "zone_no"],
        "Q5_BIW":      ["line", "zone_no"],
        "NOVA_BIW":    ["line"],
        "TCF_2":       ["shop"],
    }

    def _norm_shop_key(code: str | None) -> str:
        if not code:
            return ""
        return str(code).strip().upper()

    def _group_label(preferred_fields: list[str]) -> str:
        if not preferred_fields:
            return "Group"
        field = preferred_fields[0]
        if field == "cell":
            return "Cell"
        if field == "line":
            return "Line"
        if field == "shop":
            return "Shop"
        if field == "zone_no":
            return "Zone"
        return field.replace("_", " ").title()

    key = _norm_shop_key(shop_code)
    preferred = shop_map.get(key, ["cell", "line", "zone_no"])  # default
    group_label = _group_label(preferred)

    def _val(stn, fld):
        if fld == "cell":
            return (stn.cell or "").strip()
        if fld == "line":
            return (stn.line or "").strip()
        if fld == "zone_no":
            return (stn.zone_no or "").strip()
        if fld == "stations":
            return (stn.raw_station_id or stn.station_code or "").strip()
        return str(getattr(stn, fld, "") or "").strip()

    order = []
    groups = {}

    for stn in stations:
        vals = [_val(stn, f) for f in preferred]
        path = [v for v in vals if v]

        if not path:
            top = ""
            zone = ""
        else:
            top = path[0]
            zone = path[1] if len(path) > 1 else ""

        gkey = (top, zone)
        if gkey not in groups:
            order.append(gkey)
            groups[gkey] = []

        station_item = {
            "id": stn.id,
            "raw_id": stn.raw_station_id or stn.station_code or str(stn.id),
            "name": stn.name,
        }

        processes = []
        for proc in getattr(stn, "processes", []) or []:
            processes.append({
                "id": proc.id,
                "name": proc.name,
                "operation_count": len(getattr(proc, "operations", []) or []),
            })
        if processes:
            station_item["processes"] = processes

        groups[gkey].append(station_item)

    tops = []
    top_to_zones = {}
    for top, zone in order:
        if top not in tops:
            tops.append(top)
            top_to_zones[top] = []
        top_to_zones[top].append((zone, groups[(top, zone)]))

    hierarchy = []
    for top in tops:
        zones = []
        for zone_label, stns in top_to_zones[top]:
            zones.append({
                "label": zone_label,
                "zone_badge": _fmt_zone(zone_label) if zone_label else "",
                "stations": stns,
            })
        hierarchy.append({
            "label": top,
            "zones": zones,
        })

    return hierarchy, group_label


# ---------------------------------------------------------------------------
# Weld Shop — BIW sub-level constants and parsers
# ---------------------------------------------------------------------------

# Canonical BIW group order (as specified by the user)
_WELD_BIW_ORDER  = ["X1_BIW", "Q5_BIW", "X4_BIW", "NOVA_BIW"]

_WELD_BIW_META = {
    "X1_BIW":   {"label": "X1 BIW",   "icon": "🚗", "desc": "X1 Body-in-White Assembly"},
    "Q5_BIW":   {"label": "Q5 BIW",   "icon": "🚙", "desc": "Q5 Body-in-White Assembly"},
    "X4_BIW":   {"label": "X4 BIW",   "icon": "🛻", "desc": "X4 Body-in-White Assembly"},
    "NOVA_BIW": {"label": "NOVA BIW", "icon": "⭐", "desc": "Nova Body-in-White Assembly"},
}

# Regex: extract BIW model token from Weld station code
# e.g. WLD_X1_BIW_FR_Z1_S10 → group(1) = 'X1'
_WELD_BIW_EXTRACT = re.compile(r'^WLD_([A-Z0-9]+)_BIW', re.IGNORECASE)

# Regex: within a BIW group, extract Line and Zone
# e.g. WLD_X1_BIW_FR_Z1_S10 → line='FR', zone='1'
_WELD_BIW_SUBHIER = re.compile(r'_BIW_(?P<line>[A-Z]+)_Z(?P<zone>[0-9]+)', re.IGNORECASE)


def _get_weld_biw_subgroups(stations):
    """
    Extract the BIW model groups present in a Weld Shop station list.
    Returns a list of dicts: [{key, label, icon, desc, station_count}]
    ordered by _WELD_BIW_ORDER.
    """
    counts = {}
    for stn in stations:
        raw_id = stn.raw_station_id or stn.station_code or ""
        m = _WELD_BIW_EXTRACT.match(raw_id)
        if m:
            key = f"{m.group(1).upper()}_BIW"
            counts[key] = counts.get(key, 0) + 1

    result = []
    for key in _WELD_BIW_ORDER:
        if key in counts:
            meta = _WELD_BIW_META.get(key, {"label": key.replace("_", " "), "icon": "🔩", "desc": ""})
            result.append({
                "key":           key,
                "label":         meta["label"],
                "icon":          meta["icon"],
                "desc":          meta["desc"],
                "station_count": counts[key],
            })
    return result or None


def _parse_weld_biw_subhierarchy(stations, biw_key):
    """
    Parse Weld Shop stations belonging to one BIW group into a structured
    hierarchy. Delegates to _parse_station_hierarchy to preserve Excel row order.
    """
    model = biw_key.replace("_BIW", "").upper()   # e.g. 'X1'
    prefix = f"WLD_{model}_BIW_"                   # e.g. 'WLD_X1_BIW_'

    filtered = [
        stn for stn in stations
        if (stn.raw_station_id or "").upper().startswith(prefix)
    ]

    return _parse_station_hierarchy(filtered)


@app.route("/shop/<shop_code>")
def shop_detail(shop_code):
    with SessionLocal() as db:
        shop = _resolve_shop_record(db, shop_code)
        if not shop:
            flash(f"Shop '{shop_code}' not found.", "danger")
            return redirect(url_for("index"))
        stations = db.query(Station).filter_by(shop_id=shop.id).order_by(Station.row_order, Station.id).all()
        wis_docs = db.query(ShopWISDocument).filter_by(shop_id=shop.id).order_by(ShopWISDocument.uploaded_at.desc()).all()
        wis_workbooks = db.query(ShopWISWorkbook).filter_by(shop_id=shop.id).order_by(ShopWISWorkbook.uploaded_at.desc()).all()
        ppe_workbooks = db.query(ShopPPEWorkbook).filter_by(shop_id=shop.id).order_by(ShopPPEWorkbook.uploaded_at.desc()).all()
        is_admin = request.authorization and check_auth(request.authorization.username, request.authorization.password)

        # Weld Shop: show BIW group selector cards first
        if shop_code == "WELD_SHOP":
            subgroups = _get_weld_biw_subgroups(stations)
            if subgroups:
                return render_template(
                    "shop.html", shop=shop,
                    subgroups=subgroups, hierarchy=None, subgroup=None,
                    wis_docs=wis_docs, wis_workbooks=wis_workbooks, ppe_workbooks=ppe_workbooks, is_admin=is_admin,
                )

        hierarchy, group_label = _parse_station_hierarchy(stations, shop.shop_code)
        return render_template(
            "shop.html", shop=shop,
            hierarchy=hierarchy, group_label=group_label, subgroups=None, subgroup=None,
            wis_docs=wis_docs, wis_workbooks=wis_workbooks, ppe_workbooks=ppe_workbooks, is_admin=is_admin,
        )


@app.route("/shop/<shop_code>/<subgroup>")
def shop_subgroup(shop_code, subgroup):
    """Renders the tree for a shop sub-group (e.g. Weld BIW model)."""
    with SessionLocal() as db:
        shop = _resolve_shop_record(db, shop_code)
        if not shop:
            flash(f"Shop '{shop_code}' not found.", "danger")
            return redirect(url_for("index"))
        stations = db.query(Station).filter_by(shop_id=shop.id).order_by(Station.row_order, Station.id).all()
        wis_docs = db.query(ShopWISDocument).filter_by(shop_id=shop.id).order_by(ShopWISDocument.uploaded_at.desc()).all()
        wis_workbooks = db.query(ShopWISWorkbook).filter_by(shop_id=shop.id).order_by(ShopWISWorkbook.uploaded_at.desc()).all()
        ppe_workbooks = db.query(ShopPPEWorkbook).filter_by(shop_id=shop.id).order_by(ShopPPEWorkbook.uploaded_at.desc()).all()
        is_admin = request.authorization and check_auth(request.authorization.username, request.authorization.password)

        if shop_code == "WELD_SHOP":
            hierarchy, group_label = _parse_weld_biw_subhierarchy(stations, subgroup)
            meta = _WELD_BIW_META.get(subgroup, {})
            subgroup_label = meta.get("label", subgroup.replace("_", " "))
        else:
            hierarchy, group_label = _parse_station_hierarchy(stations, shop.shop_code)
            subgroup_label = subgroup.replace("_", " ")

        return render_template(
            "shop.html", shop=shop,
            hierarchy=hierarchy, group_label=group_label, subgroups=None, subgroup=subgroup_label,
            wis_docs=wis_docs, wis_workbooks=wis_workbooks, ppe_workbooks=ppe_workbooks, is_admin=is_admin,
        )


@app.route("/api/shop/<shop_code>/hierarchy")
def api_shop_hierarchy(shop_code):
    with SessionLocal() as db:
        shop = _resolve_shop_record(db, shop_code)
        if not shop:
            return jsonify({"error": f"Shop '{shop_code}' not found"}), 404
        stations = db.query(Station).filter_by(shop_id=shop.id).order_by(Station.row_order, Station.id).all()
        hierarchy, group_label = _parse_station_hierarchy(stations, shop.shop_code)
        return jsonify({
            "shop_code": shop.shop_code,
            "shop_name": shop.name,
            "hierarchy": hierarchy,
            "group_label": group_label,
        })

@app.route("/api/competency/<int:station_id>")
def api_competency(station_id):
    return jsonify(CompetencyEngine.score_station_readiness(station_id))


@app.route("/api/recommend/<int:station_id>")
def api_recommend(station_id):
    readiness = CompetencyEngine.score_station_readiness(station_id)
    missing_ids = [s["skill_id"] for s in readiness.get("missing_skills", [])]
    recommendations = CompetencyEngine.get_recommended_modules(missing_ids)
    return jsonify(recommendations)


@app.route("/api/stats")
def api_stats():
    return jsonify(_get_system_stats())


@app.route("/api/entity-details/<entity_type>/<entity_id>")
def api_entity_details(entity_type, entity_id):
    """
    Polymorphic endpoint resolving incoming compound string tokens (e.g., 'tool-21')
    to database integer primary keys safely to support all four parameter matrices.
    """
    try:
        if isinstance(entity_id, str) and "-" in entity_id:
            clean_id = int(entity_id.split("-")[-1])
        else:
            clean_id = int(entity_id)
    except (ValueError, IndexError, TypeError):
        logger.error(f"Malformed structural routing key received: {entity_id}")
        return jsonify({"error": f"Invalid database lookup format: {entity_id}"}), 400

    unified = build_unified_search_entity(entity_type, clean_id)
    if unified is None:
        return jsonify({"error": f"Entity contract mapping missing for: {entity_type} #{entity_id}"}), 404
    return jsonify(unified)


# ===========================================================================
# ROUTES — Admin (requires auth)
# ===========================================================================

@app.route("/admin", methods=["GET"])
@requires_auth
def admin():
    stats = _get_system_stats()
    with SessionLocal() as db:
        # Upload history from the registry (one row per file, richest data)
        upload_history = (
            db.query(UploadedFile)
            .order_by(UploadedFile.upload_time.desc())
            .limit(50)
            .all()
        )
        staging_data = (
            db.query(StagingData)
            .order_by(StagingData.created_at.desc())
            .limit(100)
            .all()
        )
        batch_summary = (
            db.query(
                StagingData.batch_id,
                StagingData.source_file,
                func.count(StagingData.id).label("total"),
                func.sum(
                    (StagingData.status == "PROCESSED").cast(_SAInteger)
                ).label("processed"),
                func.sum(
                    (StagingData.status == "FAILED").cast(_SAInteger)
                ).label("failed"),
                func.max(StagingData.created_at).label("timestamp"),
            )
            .group_by(StagingData.batch_id, StagingData.source_file)
            .order_by(func.max(StagingData.created_at).desc())
            .limit(20)
            .all()
        )
    return render_template(
        "admin.html",
        stats=stats,
        upload_history=upload_history,
        staging_data=staging_data,
        batch_summary=batch_summary,
    )


@app.route("/admin/ingest", methods=["POST"])
@requires_auth
def ingest():
    if "file" not in request.files:
        flash("No file part in request.", "danger")
        return redirect(url_for("admin"))

    file = request.files["file"]
    if not file or not file.filename:
        flash("No file selected.", "danger")
        return redirect(url_for("admin"))

    if not allowed_file(file.filename):
        flash(
            f"❌ Rejected '{file.filename}': only .xlsx and .xls files are accepted.",
            "danger",
        )
        logger.warning(f"Upload rejected — invalid extension: {file.filename}")
        return redirect(url_for("admin"))

    try:
        sha256_hash = hashlib.sha256()
        file.seek(0)
        for chunk in iter(lambda: file.read(4096), b""):
            sha256_hash.update(chunk)
        file_hash = sha256_hash.hexdigest()
        file.seek(0)  
    except Exception as hash_err:
        flash(f"❌ Failed to compute file hash: {hash_err}", "danger")
        logger.error(f"SHA-256 calculation error: {hash_err}")
        return redirect(url_for("admin"))

    with SessionLocal() as db:
        existing = db.query(UploadedFile).filter_by(file_hash=file_hash).first()
        if existing:
            # Duplicate detected — store context in session and let admin decide
            flask_session["dup_conflict"] = {
                "file_hash":     file_hash,
                "orig_name":     secure_filename(file.filename),
                "prev_filename": existing.filename,
                "prev_uploaded": str(existing.upload_time),
                "shop":          request.form.get("shop", "").strip(),
                "source_type":   request.form.get("source_type", "auto"),
                "upload_mode":   request.form.get("upload_mode", "upsert"),
                "filepath":      None,   # file not saved yet — filled below if we proceed
            }
            # Save file so it's available for reprocess action
            orig  = secure_filename(file.filename)
            stem, ext = os.path.splitext(orig)
            dup_path  = os.path.join(UPLOAD_FOLDER, f"{stem}_{_uuid.uuid4().hex[:8]}{ext}")
            file.save(dup_path)
            flask_session["dup_conflict"]["filepath"] = dup_path
            logger.info(f"Duplicate hash detected: '{file.filename}' — presenting options to admin.")
            return redirect(url_for("admin") + "?dup=1")

    original_name = secure_filename(file.filename)
    stem, ext = os.path.splitext(original_name)
    unique_name = f"{stem}_{_uuid.uuid4().hex[:8]}{ext}"
    filepath = os.path.join(UPLOAD_FOLDER, unique_name)

    try:
        file.save(filepath)
    except Exception as save_err:
        flash(f"❌ Could not save file: {save_err}", "danger")
        logger.error(f"File save error: {save_err}")
        return redirect(url_for("admin"))

    try:
        import pandas as _pd
        _pd.read_excel(filepath, nrows=1)   
    except Exception as probe_err:
        if os.path.exists(filepath):
            os.remove(filepath)   
        flash(
            f"❌ '{original_name}' could not be read as an Excel workbook: {probe_err}",
            "danger",
        )
        logger.warning(f"Excel probe failed for '{original_name}': {probe_err}")
        return redirect(url_for("admin"))

    source_type  = request.form.get("source_type", "auto")
    shop         = request.form.get("shop", "").strip()
    upload_mode  = request.form.get("upload_mode", "upsert").strip()
    if upload_mode not in {"insert_only", "update_only", "upsert", "reprocess"}:
        upload_mode = "upsert"
    try:
        logger.info(f"Admin ingestion: file='{original_name}' type='{source_type}' shop='{shop or 'N/A'}' mode='{upload_mode}'")
        stats = IngestionPipeline.ingest_excel(filepath, source_type, shop=shop, upload_mode=upload_mode)
        logger.info(f"ETL stats: {stats}")

        mapper = KnowledgeMapper()
        map_stats = mapper.run()
        logger.info(f"Mapping stats: {map_stats}")

        index_count = SearchIndexer.rebuild_index()
        logger.info(f"Search index rebuilt: {index_count} documents")

        from graph_engine import rebuild_knowledge_graph
        graph_stats = rebuild_knowledge_graph()
        logger.info(f"Knowledge Graph sync: {graph_stats}")

        with SessionLocal() as db:
            db_file = UploadedFile(
                filename=original_name,
                file_hash=file_hash,
                shop_code=shop or None,
                uploaded_by=request.authorization.username if request.authorization else None,
                upload_mode=upload_mode,
                status="ok",
            )
            db.add(db_file)
            db.commit()

        flash(
            f"✅ '{original_name}' ingested ({upload_mode}). "
            f"Inserted: {stats.get('inserted', 0)} | Updated: {stats.get('updated', 0)} | "
            f"Skipped: {stats.get('skipped', 0)} | Errors: {stats.get('errors', 0)} | "
            f"Graph Nodes: {graph_stats['nodes']}",
            "success",
        )

        # Surface any missing schema columns as a non-blocking warning
        missing_cols = stats.get("missing_columns", [])
        if missing_cols:
            # Prefer user-friendly display names when available
            try:
                from data_engine import SHOP_DATA_SCHEMA, SHOP_DATA_DISPLAY
                disp_map = {k: v for k, v in zip(SHOP_DATA_SCHEMA, SHOP_DATA_DISPLAY)}
                display_cols = [disp_map.get(c, c) for c in missing_cols]
            except Exception:
                display_cols = missing_cols
            col_list = ", ".join(display_cols)
            flash(
                f"⚠️ Missing columns detected: {col_list}\nContinuing with available data.",
                "warning",
            )

    except ValueError as ve:
        if os.path.exists(filepath):
            os.remove(filepath)
        flash(f"⚠️ Validation error: {ve}", "danger")
        logger.warning(f"Metrology or format error for '{original_name}': {ve}")
    except Exception as e:
        if os.path.exists(filepath):
            os.remove(filepath)
        flash(f"❌ Ingestion pipeline failed: {e}", "danger")
        logger.error(f"Ingestion error for '{original_name}': {e}", exc_info=True)

    return redirect(url_for("admin"))


# ---------------------------------------------------------------------------
# Duplicate-file conflict resolution
# ---------------------------------------------------------------------------

@app.route("/admin/ingest/confirm", methods=["POST"])
@requires_auth
def ingest_confirm():
    """Handle admin choice when a duplicate file hash is detected."""
    conflict = flask_session.pop("dup_conflict", None)
    if not conflict:
        flash("⚠️ No conflict data found. Please upload the file again.", "warning")
        return redirect(url_for("admin"))

    action   = request.form.get("dup_action", "skip")   # skip | reprocess
    filepath = conflict["filepath"]
    orig_name   = conflict["orig_name"]
    shop        = conflict["shop"]
    source_type = conflict["source_type"]
    upload_mode = conflict.get("upload_mode", "upsert")
    if upload_mode not in {"insert_only", "update_only", "upsert", "reprocess"}:
        upload_mode = "upsert"

    # ── SKIP: discard the saved duplicate file, do nothing
    if action == "skip":
        if filepath and os.path.exists(filepath):
            os.remove(filepath)
        flash(f"✅ Upload skipped — existing data for '{orig_name}' unchanged.", "success")
        return redirect(url_for("admin"))

    # ── REPROCESS: re-run the pipeline on the (already saved) file
    if not filepath or not os.path.exists(filepath):
        flash(f"❌ Reprocess failed — file '{orig_name}' could not be found on disk.", "danger")
        return redirect(url_for("admin"))

    # Force upsert mode for reprocess so records are updated, not blocked
    effective_mode = upload_mode if upload_mode != "reprocess" else "upsert"

    try:
        username = request.authorization.username if request.authorization else None
        logger.info(f"Reprocessing duplicate: file='{orig_name}' mode='{effective_mode}' shop='{shop or 'N/A'}'")
        stats = IngestionPipeline.ingest_excel(filepath, source_type, shop=shop, upload_mode=effective_mode)

        mapper = KnowledgeMapper()
        map_stats = mapper.run()
        SearchIndexer.rebuild_index()
        from graph_engine import rebuild_knowledge_graph
        graph_stats = rebuild_knowledge_graph()

        # Log this reprocess event as a new audit row
        with SessionLocal() as db:
            db.add(UploadedFile(
                filename=orig_name,
                file_hash=conflict["file_hash"],
                shop_code=shop or None,
                uploaded_by=username,
                upload_mode=effective_mode,
                status="reprocessed",
            ))
            db.commit()

        flash(
            f"✅ '{orig_name}' reprocessed ({effective_mode}). "
            f"Inserted: {stats.get('inserted', 0)} | Updated: {stats.get('updated', 0)} | "
            f"Skipped: {stats.get('skipped', 0)} | Errors: {stats.get('errors', 0)} | "
            f"Graph Nodes: {graph_stats['nodes']}",
            "success",
        )
        missing_cols = stats.get("missing_columns", [])
        if missing_cols:
            try:
                from data_engine import SHOP_DATA_SCHEMA, SHOP_DATA_DISPLAY
                disp_map = {k: v for k, v in zip(SHOP_DATA_SCHEMA, SHOP_DATA_DISPLAY)}
                display_cols = [disp_map.get(c, c) for c in missing_cols]
            except Exception:
                display_cols = missing_cols
            flash(f"⚠️ Missing columns detected: {', '.join(display_cols)}\nContinuing with available data.", "warning")

    except Exception as e:
        flash(f"❌ Reprocess failed: {e}", "danger")
        logger.error(f"Reprocess error for '{orig_name}': {e}", exc_info=True)

    return redirect(url_for("admin"))


@app.route("/admin/rebuild-index", methods=["POST"])
@requires_auth
def rebuild_index():
    try:
        count = SearchIndexer.rebuild_index()
        from graph_engine import rebuild_knowledge_graph
        rebuild_knowledge_graph()
        flash(f"Search index and knowledge graph rebuilt — {count} documents indexed.", "success")
    except Exception as e:
        flash(f"Index rebuild failed: {e}", "danger")
    return redirect(url_for("admin"))


@app.route("/admin/remap", methods=["POST"])
@requires_auth
def remap():
    try:
        mapper = KnowledgeMapper()
        stats = mapper.run()
        from graph_engine import rebuild_knowledge_graph
        rebuild_knowledge_graph()
        flash(f"Knowledge mapping recomputed and graph synchronized — {stats}", "success")
    except Exception as e:
        flash(f"Mapping failed: {e}", "danger")
    return redirect(url_for("admin"))


@app.route("/admin/reset-db", methods=["POST"])
@requires_auth
def reset_db():
    try:
        with SessionLocal() as db:
            db.query(StationOperationMap).delete()
            db.query(SkillStationMap).delete()
            db.query(ToolStationMap).delete()
            db.query(SkillOperationMap).delete()
            db.query(TopicSkillMap).delete()
            db.query(CompetencyMap).delete()
            db.query(GraphRelationship).delete()
            db.query(GraphEntity).delete()
            db.query(StagingData).delete()
            db.query(UploadedFile).delete()
            
            from models import Subtopic, Topic, Semester, Diploma
            from models import Skill, Tool, Operation, Process, Station, Shop
            db.query(Subtopic).delete()
            db.query(Topic).delete()
            db.query(Semester).delete()
            db.query(Diploma).delete()
            db.query(Skill).delete()
            db.query(Tool).delete()
            db.query(Operation).delete()
            db.query(Process).delete()
            db.query(Station).delete()
            db.query(Shop).delete()
            db.commit()
            logger.info("Database domain reset completed.")

        re_ingested = 0
        for fname in os.listdir(UPLOAD_FOLDER):
            if fname.lower().endswith((".xlsx", ".xls")):
                fpath = os.path.join(UPLOAD_FOLDER, fname)
                try:
                    IngestionPipeline.ingest_excel(fpath)
                    re_ingested += 1
                except Exception as fe:
                    logger.error(f"Auto re-ingest failed for {fname}: {fe}")

        mapper = KnowledgeMapper()
        mapper.run()
        SearchIndexer.rebuild_index()
        from graph_engine import rebuild_knowledge_graph
        rebuild_knowledge_graph()

        flash(f"✅ Clean database reload finished. Processed {re_ingested} layout files.", "success")
    except Exception as e:
        flash(f"❌ Reset execution aborted: {e}", "danger")
        logger.error(f"DB master clean sequence failed: {e}", exc_info=True)
    return redirect(url_for("admin"))


# ===========================================================================
# ROUTES — SHOP-LEVEL WIS / PPE WORKBOOKS
# ===========================================================================

def _persist_shop_workbook_upload(shop_code: str, file_storage, workbook_kind: str, uploaded_by: str | None, change_summary: str | None = None):
    if not file_storage or not file_storage.filename:
        flash("No file selected.", "danger")
        return False

    if not allowed_workbook_file(file_storage.filename):
        flash(f"❌ Rejected '{file_storage.filename}': only .xlsx and .xlsm files are accepted.", "danger")
        logger.warning(f"Workbook upload rejected — invalid extension: {file_storage.filename}")
        return False

    if not shop_code:
        flash("No shop selected.", "danger")
        return False

    with SessionLocal() as db:
        shop = _resolve_shop_record(db, shop_code)
        if not shop:
            flash(f"Shop '{shop_code}' not found.", "danger")
            return False

        target_dir = os.path.join(UPLOAD_FOLDER, "wis" if workbook_kind == "wis" else "ppe", "workbooks", shop.shop_code)
        os.makedirs(target_dir, exist_ok=True)

        original_name = secure_filename(file_storage.filename)
        stem, ext = os.path.splitext(original_name)
        unique_name = f"{stem}_{_uuid.uuid4().hex[:8]}{ext}"
        filepath = os.path.join(target_dir, unique_name)

        try:
            file_storage.save(filepath)
        except Exception as save_err:
            flash(f"❌ Could not save file: {save_err}", "danger")
            logger.error(f"File save error: {save_err}")
            return False

        try:
            workbook = load_workbook(filepath, read_only=True, data_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
        except Exception as parse_err:
            if os.path.exists(filepath):
                os.remove(filepath)
            flash(f"❌ Could not read workbook: {parse_err}", "danger")
            logger.error(f"Workbook parse error: {parse_err}")
            return False

        try:
            if workbook_kind == "wis":
                workbook_cls = ShopWISWorkbook
                mapping_cls = StationWISSheet
            else:
                workbook_cls = ShopPPEWorkbook
                mapping_cls = StationPPESheet

            previous_active_workbook = (
                db.query(workbook_cls)
                .filter(workbook_cls.shop_id == shop.id, workbook_cls.active.is_(True))
                .order_by(workbook_cls.uploaded_at.desc(), workbook_cls.id.desc())
                .first()
            )
            previous_mapping_lookup = {}
            if previous_active_workbook:
                previous_mappings = db.query(mapping_cls).filter_by(workbook_id=previous_active_workbook.id).all()
                previous_mapping_lookup = {mapping.sheet_name: mapping for mapping in previous_mappings}

            version_number = _get_next_workbook_version(db, workbook_cls, shop.id)
            _archive_previous_active_workbook(db, workbook_cls, shop.id, version_number)
            change_summary_value = _build_workbook_change_summary(change_summary, workbook_kind, previous_active_workbook)

            workbook_record = workbook_cls(
                shop_id=shop.id,
                file_name=original_name,
                file_path=filepath,
                sheet_count=len(sheet_names),
                version_number=version_number,
                active=True,
                archived_at=None,
                change_summary=change_summary_value,
                uploaded_by=uploaded_by,
            )
            db.add(workbook_record)
            db.flush()

            stations = db.query(Station).filter_by(shop_id=shop.id).order_by(Station.row_order, Station.id).all()
            for sheet_index, sheet_name in enumerate(sheet_names):
                matched_station = _match_station_for_sheet(sheet_name, stations)
                previous_mapping = previous_mapping_lookup.get(sheet_name)
                existing_mapping = db.query(mapping_cls).filter_by(workbook_id=workbook_record.id, sheet_name=sheet_name).first()
                if existing_mapping is None:
                    mapping_row = mapping_cls(
                        workbook_id=workbook_record.id,
                        station_id=matched_station.id if matched_station else (previous_mapping.station_id if previous_mapping is not None else None),
                        sheet_name=sheet_name,
                        sheet_index=sheet_index,
                        match_status="matched" if matched_station else (previous_mapping.match_status if previous_mapping is not None else "unmatched"),
                    )
                    _merge_sheet_mapping(mapping_row, matched_station, sheet_index, previous_mapping)
                    db.add(mapping_row)
                else:
                    _merge_sheet_mapping(existing_mapping, matched_station, sheet_index, previous_mapping)

            db.commit()
            logger.info(f"Shop {workbook_kind.upper()} workbook uploaded: shop_id={shop.id}, file='{original_name}', sheets={len(sheet_names)}")
            flash(f"✅ {workbook_kind.upper()} workbook '{original_name}' uploaded for shop '{shop.name}'.", "success")
            return True
        except Exception as db_err:
            if os.path.exists(filepath):
                os.remove(filepath)
            db.rollback()
            flash(f"❌ Could not save workbook record: {db_err}", "danger")
            logger.error(f"Database save error for workbook upload: {db_err}")
            return False


@app.route("/shop/<shop_code>/upload-wis-workbook", methods=["POST"])
@requires_auth
def upload_shop_wis_workbook(shop_code):
    file = request.files.get("file")
    uploaded_by = request.authorization.username if request.authorization else None
    change_summary = request.form.get("change_summary", "").strip() or None
    success = _persist_shop_workbook_upload(shop_code, file, "wis", uploaded_by, change_summary)
    if success:
        return redirect(url_for("shop_detail", shop_code=shop_code))
    return redirect(url_for("shop_detail", shop_code=shop_code))


@app.route("/shop/<shop_code>/upload-ppe-workbook", methods=["POST"])
@requires_auth
def upload_shop_ppe_workbook(shop_code):
    file = request.files.get("file")
    uploaded_by = request.authorization.username if request.authorization else None
    change_summary = request.form.get("change_summary", "").strip() or None
    success = _persist_shop_workbook_upload(shop_code, file, "ppe", uploaded_by, change_summary)
    if success:
        return redirect(url_for("shop_detail", shop_code=shop_code))
    return redirect(url_for("shop_detail", shop_code=shop_code))


@app.route("/admin/upload-shop-wis", methods=["POST"])
@requires_auth
def upload_shop_wis():
    file = request.files.get("file")
    shop_code = request.form.get("shop_code", "").strip() or request.form.get("shop", "").strip()
    uploaded_by = request.authorization.username if request.authorization else None
    change_summary = request.form.get("change_summary", "").strip() or None
    _persist_shop_workbook_upload(shop_code, file, "wis", uploaded_by, change_summary)
    return redirect(url_for("admin"))


@app.route("/admin/upload-station-wis", methods=["POST"])
@requires_auth
def upload_station_wis():
    flash("Station-level uploads are disabled. Use the shop page to upload WIS/PPE workbooks.", "warning")
    return redirect(url_for("admin"))


# ===========================================================================
# ROUTES — WIS Document Retrieval & Download
# ===========================================================================

@app.route("/download-shop-wis/<int:doc_id>")
@requires_auth
def download_shop_wis(doc_id):
    """Download a shop-level WIS document."""
    with SessionLocal() as db:
        doc = db.query(ShopWISDocument).filter_by(id=doc_id).first()
        if not doc:
            flash("Document not found.", "danger")
            return redirect(url_for("admin"))
        
        if not os.path.exists(doc.file_path):
            flash("Document file not found on disk.", "danger")
            return redirect(url_for("admin"))
        
        try:
            return send_file(
                doc.file_path,
                as_attachment=True,
                download_name=doc.file_name,
                mimetype="application/octet-stream"
            )
        except Exception as e:
            logger.error(f"Download failed for doc {doc_id}: {e}")
            flash(f"Download failed: {e}", "danger")
            return redirect(url_for("admin"))


@app.route("/view-shop-wis/<int:doc_id>")
def view_shop_wis(doc_id):
    """View a shop-level WIS document in browser (for PDF/media)."""
    with SessionLocal() as db:
        doc = db.query(ShopWISDocument).filter_by(id=doc_id).first()
        if not doc:
            return jsonify({"error": "Document not found"}), 404
        
        if not os.path.exists(doc.file_path):
            return jsonify({"error": "Document file not found"}), 404
        
        try:
            ext = os.path.splitext(doc.file_name.lower())[1]
            mimetype_map = {
                ".pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
                ".ppt": "application/vnd.ms-powerpoint",
                ".pdf": "application/pdf",
                ".png": "image/png",
                ".jpg": "image/jpeg",
                ".jpeg": "image/jpeg",
            }
            mimetype = mimetype_map.get(ext, "application/octet-stream")
            return send_file(doc.file_path, mimetype=mimetype)
        except Exception as e:
            logger.error(f"View failed for doc {doc_id}: {e}")
            return jsonify({"error": "View failed"}), 500


@app.route("/download-station-wis/<int:doc_id>")
@requires_auth
def download_station_wis(doc_id):
    """Download a station-level WIS document."""
    with SessionLocal() as db:
        doc = db.query(StationWISDocument).filter_by(id=doc_id).first()
        if not doc:
            flash("Document not found.", "danger")
            return redirect(url_for("admin"))
        
        if not os.path.exists(doc.file_path):
            flash("Document file not found on disk.", "danger")
            return redirect(url_for("admin"))
        
        try:
            return send_file(
                doc.file_path,
                as_attachment=True,
                download_name=doc.file_name,
                mimetype="application/octet-stream"
            )
        except Exception as e:
            logger.error(f"Download failed for doc {doc_id}: {e}")
            flash(f"Download failed: {e}", "danger")
            return redirect(url_for("admin"))


@app.route("/view-station-wis/<int:doc_id>")
def view_station_wis(doc_id):
    """View a station-level WIS document in browser (for PDF/media)."""
    with SessionLocal() as db:
        doc = db.query(StationWISDocument).filter_by(id=doc_id).first()
        if not doc:
            return jsonify({"error": "Document not found"}), 404
        
        if not os.path.exists(doc.file_path):
            return jsonify({"error": "Document file not found"}), 404
        
        try:
            ext = os.path.splitext(doc.file_name.lower())[1]
            mimetype_map = {
                ".pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
                ".ppt": "application/vnd.ms-powerpoint",
                ".pdf": "application/pdf",
                ".png": "image/png",
                ".jpg": "image/jpeg",
                ".jpeg": "image/jpeg",
            }
            mimetype = mimetype_map.get(ext, "application/octet-stream")
            return send_file(doc.file_path, mimetype=mimetype)
        except Exception as e:
            logger.error(f"View failed for doc {doc_id}: {e}")
            return jsonify({"error": "View failed"}), 500


# ===========================================================================
# ROUTES — WIS Document Viewer Pages
# ===========================================================================

@app.route("/shop/wis/<int:doc_id>")
def shop_wis_viewer(doc_id):
    """Display shop WIS document viewer page with metadata."""
    with SessionLocal() as db:
        doc = db.query(ShopWISDocument).filter_by(id=doc_id).first()
        if not doc:
            flash("WIS document not found.", "danger")
            return redirect(url_for("index"))
        
        shop = doc.shop
        return render_template(
            "wis-viewer.html",
            doc_id=doc.id,
            file_name=doc.file_name,
            uploaded_at=doc.uploaded_at,
            uploaded_by=doc.uploaded_by,
            entity_type="shop",
            entity_name=shop.name,
            entity_code=shop.shop_code,
            view_url=url_for("view_shop_wis", doc_id=doc.id),
            download_url=url_for("download_shop_wis", doc_id=doc.id),
        )


@app.route("/station/wis/<int:doc_id>")
def station_wis_viewer(doc_id):
    """Display station WIS document viewer page with metadata."""
    with SessionLocal() as db:
        doc = db.query(StationWISDocument).filter_by(id=doc_id).first()
        if not doc:
            flash("WIS document not found.", "danger")
            return redirect(url_for("index"))
        
        station = doc.station
        shop = station.shop
        return render_template(
            "wis-viewer.html",
            doc_id=doc.id,
            file_name=doc.file_name,
            uploaded_at=doc.uploaded_at,
            uploaded_by=doc.uploaded_by,
            entity_type="station",
            entity_name=station.station_code,
            entity_code=station.station_code,
            station_id=station.id,
            shop_code=shop.shop_code if shop else "N/A",
            shop_name=shop.name if shop else "N/A",
            view_url=url_for("view_station_wis", doc_id=doc.id),
            download_url=url_for("download_station_wis", doc_id=doc.id),
        )


def _wal_checkpoint():
    try:
        with engine.connect() as conn:
            conn.execute(text("PRAGMA wal_checkpoint(TRUNCATE)"))
            logger.info("WAL checkpoint finalized safely on shutdown sequence.")
    except Exception as e:
        logger.warning(f"WAL checkpoint failure: {e}")



init_db()
_ensure_workbook_versioning_schema()

atexit.register(_wal_checkpoint)

if __name__ == "__main__":
    init_db()
    _ensure_workbook_versioning_schema()
    logger.info(f"Starting IIK-CME production server on port 5000. DEBUG={DEBUG_MODE}")
    app.run(host="0.0.0.0", port=5000, debug=DEBUG_MODE, use_reloader=DEBUG_MODE)